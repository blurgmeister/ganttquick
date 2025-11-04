#!/usr/bin/env python3
"""
Test error handling for Excel import
"""

from openpyxl import Workbook
from excel_import import import_from_excel, ExcelImportError

def test_missing_tabs():
    """Test that import fails appropriately when required tabs are missing."""

    print("\nTesting missing tabs validation...")

    # Create a workbook with only some tabs
    wb = Workbook()
    ws = wb.active
    ws.title = "Gantt Chart"

    # Missing "Project Info" and "Work Schedules"
    test_file = "test_missing_tabs.xlsx"
    wb.save(test_file)
    wb.close()

    try:
        import_from_excel(test_file)
        print("   ERROR: Should have failed with missing tabs!")
        return False
    except ExcelImportError as e:
        error_msg = str(e)
        print(f"   Correctly caught error: {error_msg}")
        assert "Missing required tabs" in error_msg
        assert "Project Info" in error_msg
        assert "Work Schedules" in error_msg
        print("   Test passed!")
        return True

def test_optional_holiday_tab():
    """Test that import works when Holiday Schedule tab is missing."""

    print("\nTesting optional Holiday Schedule tab...")

    # Create a minimal valid workbook without Holiday Schedule
    wb = Workbook()

    # Gantt Chart tab
    ws = wb.active
    ws.title = "Gantt Chart"
    ws['A1'] = "Task Name"
    ws['A3'] = "Sample Task"
    ws['C3'] = "Alice"
    ws['D3'] = 5

    # Project Info tab
    info = wb.create_sheet("Project Info")
    info['A1'] = "Project Name:"
    info['B1'] = "Test Project"
    info['A2'] = "Start Date:"
    info['B2'] = "2025-01-15"

    # Work Schedules tab
    work = wb.create_sheet("Work Schedules")
    work['A1'] = "Employee Name"
    work['A2'] = "Alice"
    work['B2'] = "X"  # Monday
    work['C2'] = "X"  # Tuesday

    test_file = "test_no_holidays.xlsx"
    wb.save(test_file)
    wb.close()

    try:
        data = import_from_excel(test_file)
        print(f"   Import successful without Holiday Schedule tab")
        print(f"   Global holidays: {data['project_info']['global_holidays']}")
        print(f"   Employee holidays: {data['employees'][0]['holidays']}")
        assert data['project_info']['global_holidays'] == []
        assert data['employees'][0]['holidays'] == []
        print("   Test passed!")
        return True
    except Exception as e:
        print(f"   ERROR: {e}")
        return False

if __name__ == '__main__':
    test1 = test_missing_tabs()
    test2 = test_optional_holiday_tab()

    if test1 and test2:
        print("\n" + "=" * 60)
        print("ALL ERROR HANDLING TESTS PASSED!")
        print("=" * 60)
        exit(0)
    else:
        print("\nSOME TESTS FAILED!")
        exit(1)
