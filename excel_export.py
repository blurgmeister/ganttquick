from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime, timedelta
from models import Project


def export_to_excel(project: Project, filename: str = "gantt_chart.xlsx"):
    """Export the project Gantt chart to an Excel file."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Gantt Chart"

    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    task_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    holiday_fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Get date range
    start_date, end_date = project.get_date_range()
    if not end_date:
        end_date = start_date

    # Create date list (all calendar days from start to end)
    date_list = []
    current = start_date
    while current <= end_date:
        date_list.append(current)
        current += timedelta(days=1)

    # Header row 1: Fixed columns
    headers = ["Task Name", "Depends On", "Assigned To", "Estimated Duration", "Availability (%)", "Contingency (%)", "Actual Duration", "Custom Start Date", "Start Date", "End Date"]
    col_offset = len(headers) + 1  # +1 for spacing

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Header row 1: Date columns
    for idx, date in enumerate(date_list):
        col = col_offset + idx
        cell = ws.cell(row=1, column=col)
        cell.value = date.strftime("%m/%d")
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", text_rotation=90)
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = 3

    # Header row 2: Day of week
    for idx, date in enumerate(date_list):
        col = col_offset + idx
        cell = ws.cell(row=2, column=col)
        cell.value = date.strftime("%a")[0]  # M, T, W, T, F, S, S
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Set column widths for fixed columns
    ws.column_dimensions['A'].width = 30  # Task Name
    ws.column_dimensions['B'].width = 20  # Depends On
    ws.column_dimensions['C'].width = 20  # Assigned To
    ws.column_dimensions['D'].width = 15  # Estimated Duration
    ws.column_dimensions['E'].width = 15  # Availability (%)
    ws.column_dimensions['F'].width = 15  # Contingency (%)
    ws.column_dimensions['G'].width = 15  # Actual Duration
    ws.column_dimensions['H'].width = 15  # Custom Start Date
    ws.column_dimensions['I'].width = 12  # Start Date
    ws.column_dimensions['J'].width = 12  # End Date

    # Data rows
    for task_idx, task in enumerate(project.tasks, start=3):
        # Fixed columns
        ws.cell(row=task_idx, column=1).value = task.name
        ws.cell(row=task_idx, column=2).value = task.dependency if task.dependency else ""
        ws.cell(row=task_idx, column=3).value = task.assigned_to
        ws.cell(row=task_idx, column=4).value = task.estimated_duration
        ws.cell(row=task_idx, column=5).value = task.availability
        ws.cell(row=task_idx, column=6).value = task.contingency_margin
        ws.cell(row=task_idx, column=7).value = task.actual_duration
        ws.cell(row=task_idx, column=8).value = task.custom_start_date.strftime("%Y-%m-%d") if task.custom_start_date else ""
        ws.cell(row=task_idx, column=9).value = task.start_date.strftime("%Y-%m-%d") if task.start_date else ""
        ws.cell(row=task_idx, column=10).value = task.end_date.strftime("%Y-%m-%d") if task.end_date else ""

        # Apply borders to fixed columns
        for col in range(1, 11):
            ws.cell(row=task_idx, column=col).border = border
            ws.cell(row=task_idx, column=col).alignment = Alignment(vertical="center")

        # Date columns - highlight working days and holidays
        working_dates_set = {d.strftime("%Y-%m-%d") for d in task.working_dates}
        holiday_dates_set = {d.strftime("%Y-%m-%d") for d in task.holiday_dates}
        for idx, date in enumerate(date_list):
            col = col_offset + idx
            cell = ws.cell(row=task_idx, column=col)
            cell.border = border

            date_str = date.strftime("%Y-%m-%d")
            if date_str in working_dates_set:
                cell.fill = task_fill
                cell.value = 1
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif date_str in holiday_dates_set:
                cell.fill = holiday_fill
                cell.value = 0
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add project info sheet
    info_ws = wb.create_sheet("Project Info")
    info_ws['A1'] = "Project Name:"
    info_ws['B1'] = project.name
    info_ws['A2'] = "Start Date:"
    info_ws['B2'] = project.start_date.strftime("%Y-%m-%d")
    info_ws['A3'] = "End Date:"
    info_ws['B3'] = end_date.strftime("%Y-%m-%d") if end_date else ""
    info_ws['A4'] = "Total Duration (days):"
    info_ws['B4'] = (end_date - project.start_date).days + 1 if end_date else 0

    # Bold the labels
    for row in range(1, 5):
        info_ws.cell(row=row, column=1).font = Font(bold=True)

    info_ws.column_dimensions['A'].width = 25
    info_ws.column_dimensions['B'].width = 25

    # Add Work Schedules sheet
    work_ws = wb.create_sheet("Work Schedules")
    work_ws['A1'] = "Employee Name"
    work_ws['A1'].font = Font(bold=True)
    work_ws['A1'].fill = header_fill
    work_ws['A1'].font = header_font
    work_ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    work_ws['A1'].border = border

    # Days of week headers
    days_of_week = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    for col_idx, day in enumerate(days_of_week, start=2):
        cell = work_ws.cell(row=1, column=col_idx)
        cell.value = day
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Add employee work patterns
    for row_idx, (emp_name, employee) in enumerate(sorted(project.employees.items()), start=2):
        work_ws.cell(row=row_idx, column=1).value = emp_name
        work_ws.cell(row=row_idx, column=1).border = border
        work_ws.cell(row=row_idx, column=1).alignment = Alignment(vertical="center")

        # Mark working days with "X"
        for day_num in range(7):  # 0=Monday, 6=Sunday
            col = day_num + 2
            cell = work_ws.cell(row=row_idx, column=col)
            if day_num in employee.work_pattern:
                cell.value = "X"
                cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    # Set column widths for Work Schedules
    work_ws.column_dimensions['A'].width = 20
    for col in range(2, 9):
        work_ws.column_dimensions[work_ws.cell(row=1, column=col).column_letter].width = 12

    # Add Holiday Schedule sheet
    holiday_ws = wb.create_sheet("Holiday Schedule")
    holiday_ws['A1'] = "Employee Name"
    holiday_ws['A1'].font = Font(bold=True)
    holiday_ws['A1'].fill = header_fill
    holiday_ws['A1'].font = header_font
    holiday_ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    holiday_ws['A1'].border = border

    holiday_ws['B1'] = "Holiday Dates"
    holiday_ws['B1'].font = Font(bold=True)
    holiday_ws['B1'].fill = header_fill
    holiday_ws['B1'].font = header_font
    holiday_ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
    holiday_ws['B1'].border = border

    # Add global holidays first
    holiday_ws['A2'] = "GLOBAL"
    holiday_ws['A2'].font = Font(bold=True)
    holiday_ws['A2'].border = border
    holiday_ws['A2'].alignment = Alignment(vertical="center")

    if project.global_holidays:
        sorted_global_holidays = sorted(project.global_holidays)
        holiday_ws.cell(row=2, column=2).value = ", ".join(sorted_global_holidays)
    else:
        holiday_ws.cell(row=2, column=2).value = "None"
    holiday_ws.cell(row=2, column=2).border = border
    holiday_ws.cell(row=2, column=2).alignment = Alignment(vertical="center", wrap_text=True)

    # Add employee-specific holidays
    row_idx = 3
    for emp_name, employee in sorted(project.employees.items()):
        holiday_ws.cell(row=row_idx, column=1).value = emp_name
        holiday_ws.cell(row=row_idx, column=1).border = border
        holiday_ws.cell(row=row_idx, column=1).alignment = Alignment(vertical="center")

        if employee.holidays:
            sorted_holidays = sorted(employee.holidays)
            holiday_ws.cell(row=row_idx, column=2).value = ", ".join(sorted_holidays)
        else:
            holiday_ws.cell(row=row_idx, column=2).value = "None"
        holiday_ws.cell(row=row_idx, column=2).border = border
        holiday_ws.cell(row=row_idx, column=2).alignment = Alignment(vertical="center", wrap_text=True)
        row_idx += 1

    # Set column widths for Holiday Schedule
    holiday_ws.column_dimensions['A'].width = 20
    holiday_ws.column_dimensions['B'].width = 60

    # Save the file
    wb.save(filename)
    return filename
