from openpyxl import load_workbook
from datetime import datetime, timedelta
import re


class ExcelImportError(Exception):
    """Custom exception for Excel import errors."""
    pass


def _parse_date_flexible(date_str):
    """
    Parse date from either dd/mm/yyyy or YYYY-MM-DD format.
    Returns date in YYYY-MM-DD format.
    """
    if not date_str:
        return None

    date_str = str(date_str).strip()
    if not date_str:
        return None

    # Try dd/mm/yyyy format first (new format)
    try:
        parsed_date = datetime.strptime(date_str, '%d/%m/%Y')
        return parsed_date.strftime('%Y-%m-%d')
    except ValueError:
        pass

    # Try YYYY-MM-DD format (backwards compatibility)
    try:
        parsed_date = datetime.strptime(date_str, '%Y-%m-%d')
        return parsed_date.strftime('%Y-%m-%d')
    except ValueError:
        pass

    # If both fail, raise error
    raise ValueError(f"Invalid date format: {date_str}. Expected dd/mm/yyyy or YYYY-MM-DD")


def _parse_date_ranges(date_input):
    """
    Parse date ranges and individual dates from a string.

    Accepts formats:
    - Individual dates: dd/mm/yyyy
    - Date ranges: dd/mm/yyyy-dd/mm/yyyy or dd/mm/yyyy - dd/mm/yyyy
    - Multiple entries separated by commas

    Returns a list of dates in YYYY-MM-DD format.
    """
    if not date_input or not str(date_input).strip():
        return []

    date_input = str(date_input).strip()
    result_dates = []

    # Split by commas
    entries = [entry.strip() for entry in date_input.split(',')]

    for entry in entries:
        if not entry:
            continue

        # Check if this is a date range (contains a dash)
        # Pattern: dd/mm/yyyy - dd/mm/yyyy or dd/mm/yyyy-dd/mm/yyyy
        range_pattern = r'(\d{1,2}/\d{1,2}/\d{4})\s*-\s*(\d{1,2}/\d{1,2}/\d{4})'
        match = re.match(range_pattern, entry)

        if match:
            # This is a date range
            start_str = match.group(1)
            end_str = match.group(2)

            try:
                # Parse start and end dates
                start_date = datetime.strptime(start_str, '%d/%m/%Y')
                end_date = datetime.strptime(end_str, '%d/%m/%Y')

                if start_date > end_date:
                    raise ValueError(f"Start date {start_str} is after end date {end_str}")

                # Generate all dates in the range
                current_date = start_date
                while current_date <= end_date:
                    result_dates.append(current_date.strftime('%Y-%m-%d'))
                    current_date += timedelta(days=1)
            except ValueError as e:
                # Skip invalid date ranges but continue processing
                pass
        else:
            # This is a single date
            try:
                parsed_date = _parse_date_flexible(entry)
                if parsed_date:
                    result_dates.append(parsed_date)
            except ValueError:
                # Skip invalid dates but continue processing
                pass

    return result_dates


def import_from_excel(filepath):
    """
    Import project data from an Excel file.

    Returns a dictionary with:
    - project_info: {name, start_date, global_holidays}
    - employees: [{name, work_pattern, holidays}, ...]
    - tasks: [{name, dependency, assigned_to, estimated_duration, availability, contingency_margin, custom_start_date}, ...]
    """
    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception as e:
        raise ExcelImportError(f"Failed to open Excel file: {str(e)}")

    # Validate required tabs exist
    required_tabs = ['Gantt Chart', 'Project Info', 'Work Schedules']
    missing_tabs = [tab for tab in required_tabs if tab not in wb.sheetnames]

    if missing_tabs:
        raise ExcelImportError(f"Missing required tabs: {', '.join(missing_tabs)}")

    # Check if Holiday Schedule tab exists
    has_holiday_tab = 'Holiday Schedule' in wb.sheetnames

    # Extract data
    project_info = _extract_project_info(wb['Project Info'])
    employees = _extract_employees(wb['Work Schedules'])

    if has_holiday_tab:
        _extract_holidays(wb['Holiday Schedule'], project_info, employees)

    tasks = _extract_tasks(wb['Gantt Chart'])

    wb.close()

    return {
        'project_info': project_info,
        'employees': employees,
        'tasks': tasks
    }


def _extract_project_info(ws):
    """Extract project information from Project Info tab."""
    project_info = {
        'name': None,
        'start_date': None,
        'global_holidays': []
    }

    # Read project name (A1:B1)
    if ws['A1'].value and str(ws['A1'].value).strip().lower().startswith('project name'):
        project_info['name'] = str(ws['B1'].value).strip() if ws['B1'].value else 'Imported Project'

    # Read start date (A2:B2)
    if ws['A2'].value and str(ws['A2'].value).strip().lower().startswith('start date'):
        start_date_value = ws['B2'].value
        if start_date_value:
            if isinstance(start_date_value, datetime):
                project_info['start_date'] = start_date_value.strftime('%Y-%m-%d')
            else:
                # Try to parse as string
                try:
                    project_info['start_date'] = _parse_date_flexible(str(start_date_value))
                except ValueError as e:
                    raise ExcelImportError(f"Invalid start date: {e}")

    if not project_info['name'] or not project_info['start_date']:
        raise ExcelImportError("Project Info tab must contain Project Name and Start Date")

    return project_info


def _extract_employees(ws):
    """Extract employee work schedules from Work Schedules tab."""
    employees = []

    # Header should be in row 1
    # Columns: A=Employee Name, B-H=Mon-Sun

    # Start from row 2 (first data row)
    row = 2
    while True:
        emp_name = ws.cell(row=row, column=1).value
        if not emp_name:
            break

        emp_name = str(emp_name).strip()

        # Extract work pattern (columns 2-8 for Mon-Sun)
        work_pattern = []
        for day_num in range(7):  # 0=Monday, 6=Sunday
            col = day_num + 2
            cell_value = ws.cell(row=row, column=col).value
            if cell_value and str(cell_value).strip().upper() == 'X':
                work_pattern.append(day_num)

        employees.append({
            'name': emp_name,
            'work_pattern': work_pattern,
            'holidays': []  # Will be filled in if Holiday Schedule exists
        })

        row += 1

    if not employees:
        raise ExcelImportError("No employees found in Work Schedules tab")

    return employees


def _extract_holidays(ws, project_info, employees):
    """Extract holiday information from Holiday Schedule tab."""

    # Create a dict for quick employee lookup
    emp_dict = {emp['name']: emp for emp in employees}

    # Start from row 2 (first data row)
    row = 2
    while True:
        emp_name = ws.cell(row=row, column=1).value
        if not emp_name:
            break

        emp_name = str(emp_name).strip()
        holidays_str = ws.cell(row=row, column=2).value

        # Parse holidays (supports individual dates and date ranges)
        holidays = []
        if holidays_str and str(holidays_str).strip().upper() != 'NONE':
            holidays = _parse_date_ranges(holidays_str)

        # Assign to global or employee-specific
        if emp_name.upper() == 'GLOBAL':
            project_info['global_holidays'] = holidays
        elif emp_name in emp_dict:
            emp_dict[emp_name]['holidays'] = holidays

        row += 1


def _extract_tasks(ws):
    """Extract tasks from Gantt Chart tab."""
    tasks = []

    # Header is in rows 1-2
    # Data starts from row 3
    # Columns: A=Task Name, B=Depends On, C=Assigned To, D=Estimated Duration,
    #          E=Availability (%), F=Contingency (%), G=Actual Duration,
    #          H=Custom Start Date, I=Start Date, J=End Date

    row = 3
    while True:
        task_name = ws.cell(row=row, column=1).value  # Column A
        if not task_name:
            break

        task_name = str(task_name).strip()

        # Extract task data
        dependency = ws.cell(row=row, column=2).value  # Column B
        if dependency:
            dependency = str(dependency).strip()
            if not dependency:
                dependency = None
        else:
            dependency = None

        assigned_to = ws.cell(row=row, column=3).value  # Column C
        if not assigned_to:
            raise ExcelImportError(f"Task '{task_name}' has no assigned employee")
        assigned_to = str(assigned_to).strip()

        estimated_duration = ws.cell(row=row, column=4).value  # Column D
        if not estimated_duration:
            raise ExcelImportError(f"Task '{task_name}' has no estimated duration")
        try:
            estimated_duration = int(estimated_duration)
        except (ValueError, TypeError):
            raise ExcelImportError(f"Task '{task_name}' has invalid estimated duration: {estimated_duration}")

        availability = ws.cell(row=row, column=5).value  # Column E
        if availability is None or availability == '':
            availability = 100
        try:
            availability = int(availability)
        except (ValueError, TypeError):
            raise ExcelImportError(f"Task '{task_name}' has invalid availability: {availability}")

        contingency_margin = ws.cell(row=row, column=6).value  # Column F
        if contingency_margin is None or contingency_margin == '':
            contingency_margin = 0
        try:
            contingency_margin = int(contingency_margin)
        except (ValueError, TypeError):
            raise ExcelImportError(f"Task '{task_name}' has invalid contingency margin: {contingency_margin}")

        custom_start_date = ws.cell(row=row, column=8).value  # Column H
        if custom_start_date:
            if isinstance(custom_start_date, datetime):
                custom_start_date = custom_start_date.strftime('%Y-%m-%d')
            else:
                custom_start_date_str = str(custom_start_date).strip()
                if custom_start_date_str:
                    try:
                        custom_start_date = _parse_date_flexible(custom_start_date_str)
                    except ValueError:
                        custom_start_date = None
                else:
                    custom_start_date = None
        else:
            custom_start_date = None

        tasks.append({
            'name': task_name,
            'dependency': dependency,
            'assigned_to': assigned_to,
            'estimated_duration': estimated_duration,
            'availability': availability,
            'contingency_margin': contingency_margin,
            'custom_start_date': custom_start_date
        })

        row += 1

    if not tasks:
        raise ExcelImportError("No tasks found in Gantt Chart tab")

    return tasks
